import os
import sys
import subprocess
import webbrowser
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import openpyxl
from PIL import Image

from excel_manager import (
    get_or_create_excel, 
    read_config_from_excel, 
    read_open_excel_via_com,
    write_files_to_seznam, 
    update_autel_config,
    safe_save_workbook,
    safe_load_workbook,
    DEFAULT_EXCEL_NAME
)
from scanner import scan_directory
from pdf_processor import process_pdfs_from_excel

APP_VERSION = "v1.2.0"
APP_AUTHOR = "KBK"
GIT_REPOSITORY_URL = "https://github.com/KBK/Utilita-Dokumentace-Kvality"
MANUAL_PDF_NAME = "Uzivatelska_Prirucka_Utilita_Dokumentace_Kvality.pdf"

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

PALETTE_COLORS = [
    ("#000000", "Černá"),
    ("#FFFFFF", "Bílá"),
    ("#FF0000", "Červená"),
    ("#800000", "Vínová"),
    ("#0000FF", "Modrá"),
    ("#000080", "Tmavě modrá"),
    ("#00FF00", "Zelená"),
    ("#008000", "Tmavě zelená"),
    ("#FFFF00", "Žlutá"),
    ("#FFA500", "Oranžová"),
    ("#800080", "Fialová"),
    ("#00FFFF", "Světle modrá"),
    ("#808080", "Šedá"),
    ("#D3D3D3", "Světle šedá"),
    ("#A52A2A", "Hnědá"),
    ("TRANSPARENT", "Bez pozadí / Průhledná")
]

class ManualWindow(ctk.CTkToplevel):
    """
    Samostatné okno s kompletním uživatelským návodem přímo v aplikaci.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title(f"Uživatelská příručka & Návod - AUTEL ({APP_VERSION})")
        self.geometry("780x640")
        self.attributes("-topmost", True)
        
        # Hlavička návodu
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(padx=20, pady=(15, 5), fill="x")
        
        ctk.CTkLabel(
            header_frame, 
            text="📖 Uživatelská Příručka & Návod k Použití", 
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            header_frame, 
            text=f"AUTEL - Utilita pro Tvorbu Dokumentace Kvality ({APP_VERSION})  |  Autor: {APP_AUTHOR}", 
            font=ctk.CTkFont(size=11, slant="italic"),
            text_color="#AAAAAA"
        ).pack(anchor="w", pady=(2, 0))
        
        # Textové pole s návodem
        self.txt_manual = ctk.CTkTextbox(self, font=ctk.CTkFont(family="Consolas", size=11), corner_radius=8)
        self.txt_manual.pack(padx=20, pady=10, fill="both", expand=True)
        
        manual_content = (
            "========================================================================\n"
            f"   AUTEL, a.s. - UTILITA PRO TVORBU DOKUMENTACE KVALITY ({APP_VERSION})\n"
            f"   Autor: {APP_AUTHOR} | Datum: 2026\n"
            f"   Git Repozitář: {GIT_REPOSITORY_URL}\n"
            "========================================================================\n\n"
            "1. ÚČEL APLIKACE:\n"
            "   Utilita automatizuje zpracování, hromadné razítkování, přečíslování a\n"
            "   kompletaci PDF dokumentů kvality (Quality Documentation) podle podadresářové struktury.\n\n"
            "2. REGULA A PRINCIP ČÍSLOVÁNÍ:\n"
            "   - Kód podadresáře se sestavuje z čísel složek (např. 1.1, 1.2, 2.0, 4.1, 4.2).\n"
            "   - Soubory v daném adresáři dostávají dvojmístné pořadí (.01, .02, .03...).\n"
            "   - Výsledný kód: 1.1.01, 1.2.01 až 1.2.07, 2.0.01, 4.1.01, 4.1.02 atd.\n"
            "   - Funguje automaticky i pro 5 a více úrovní vnořených podadresářů bez duplikace.\n\n"
            "3. NÁVOD K POUŽITÍ:\n"
            "   1) Vyberte Zdrojový adresář se soubory PDF nebo podklady.\n"
            "   2) Klikněte na 'Vygenerovat / Obnovit Excel' -> Vytvoří se sešit Dokumentace_Kvality.xlsx.\n"
            "   3) Nastavte pravidla (otáčení stránek, zachování struktury složek, sloučení do PDF).\n"
            "   4) Na záložce 'Formát Razítka' zvolte velikost, průhlednost, styl fontu a barvu razítka.\n"
            "   5) Všechny změny v GUI se OKAMŽITĚ automaticky ukládají do záložky AUTEL v Excelu!\n"
            "   6) Klikněte na 'Spustit razítkování podle Excelu' -> Aplikace orazítkuje PDF v pravém\n"
            "      horním rohu a vytvoří Master PDF se strukturovanými záložkami (TOC).\n\n"
            "4. NAVIGACE A ZÁLOŽKY V MASTER PDF:\n"
            "   - Při spuštění kompletace do jednoho PDF se v Master PDF vytvoří strom záložek.\n"
            "   - Záložky odkazují přesně na souřadnice razítka s označením v pravém horním rohu.\n\n"
            "========================================================================\n"
            f"   Autor: {APP_AUTHOR}  |  Git: {GIT_REPOSITORY_URL}\n"
            "========================================================================\n"
        )
        self.txt_manual.insert("1.0", manual_content)
        self.txt_manual.configure(state="disabled")
        
        # Tlačítka dole
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(padx=20, pady=(5, 15), fill="x")
        
        ctk.CTkButton(
            btn_frame, 
            text="🌐 Otevřít Git Repozitář", 
            command=lambda: webbrowser.open(GIT_REPOSITORY_URL),
            fg_color="#24292e",
            hover_color="#1b1f23",
            height=36
        ).pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(
            btn_frame, 
            text="📄 Otevřít PDF Příručku", 
            command=self.parent.open_pdf_manual,
            fg_color="#1F497D",
            hover_color="#143054",
            height=36
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            btn_frame, 
            text="Zavřít", 
            command=self.destroy,
            width=100,
            height=36
        ).pack(side="right", padx=(10, 0))

class GraphicalColorPicker(ctk.CTkFrame):
    """
    Grafický vzorník barev s 16 tlačítky a vizuálním náhledem vybrané barvy.
    """
    def __init__(self, parent, title="Barva", default_hex="#000000", is_bg=False, on_change=None):
        super().__init__(parent, fg_color="transparent")
        
        self.is_bg = is_bg
        self.on_change = on_change
        self.selected_hex = default_hex.upper() if default_hex else "#000000"
        if is_bg and (self.selected_hex in ["NONE", "TRANSPARENT", "0", "#FFFFFF00"]):
            self.selected_hex = "TRANSPARENT"
            
        self.buttons = {}
        
        self.lbl_title = ctk.CTkLabel(self, text=title, font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_title.pack(anchor="w", pady=(0, 4))
        
        self.content_frame = ctk.CTkFrame(self, fg_color="#242424", corner_radius=8, border_width=1, border_color="#3a3a3a")
        self.content_frame.pack(fill="x", pady=2)
        
        self.preview_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.preview_frame.pack(anchor="w", padx=10, pady=(8, 4))
        
        self.box_preview = ctk.CTkFrame(self.preview_frame, width=22, height=22, corner_radius=4, border_width=1, border_color="#ffffff")
        self.box_preview.pack(side="left", padx=(0, 8))
        
        self.lbl_preview_name = ctk.CTkLabel(self.preview_frame, text="", font=ctk.CTkFont(size=11, weight="bold"))
        self.lbl_preview_name.pack(side="left")
        
        self.grid_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.grid_frame.pack(padx=8, pady=(4, 8))
        
        for idx, (hex_code, name) in enumerate(PALETTE_COLORS):
            row = idx // 8
            col = idx % 8
            
            btn_fg = "#2b2b2b" if hex_code == "TRANSPARENT" else hex_code
            btn_text = "✕" if hex_code == "TRANSPARENT" else ""
            btn_text_color = "white" if hex_code in ["#000000", "#800000", "#000080", "#008000", "#800080", "TRANSPARENT"] else "black"
            
            btn = ctk.CTkButton(
                self.grid_frame,
                text=btn_text,
                text_color=btn_text_color,
                fg_color=btn_fg,
                hover_color=btn_fg,
                width=32,
                height=32,
                corner_radius=6,
                border_width=2,
                border_color="#444444",
                command=lambda h=hex_code: self.select_color(h)
            )
            btn.grid(row=row, column=col, padx=3, pady=3)
            self.buttons[hex_code] = btn
            
        self.update_ui()

    def select_color(self, hex_code):
        self.selected_hex = hex_code
        self.update_ui()
        if self.on_change:
            self.on_change()

    def update_ui(self):
        for hex_code, btn in self.buttons.items():
            if hex_code.upper() == self.selected_hex.upper():
                btn.configure(border_color="#00D2FF", border_width=3)
            else:
                btn.configure(border_color="#444444", border_width=2)
                
        name = "Neznámá"
        for h, n in PALETTE_COLORS:
            if h.upper() == self.selected_hex.upper():
                name = n
                break
                
        if self.selected_hex == "TRANSPARENT":
            self.box_preview.configure(fg_color="#2b2b2b")
            self.lbl_preview_name.configure(text="Bez pozadí (Průhledné)")
        else:
            self.box_preview.configure(fg_color=self.selected_hex)
            self.lbl_preview_name.configure(text=f"{name} ({self.selected_hex})")

    def get_hex(self):
        return self.selected_hex

    def set_hex(self, hex_code):
        hex_code = str(hex_code).strip().upper()
        if self.is_bg and (hex_code in ["NONE", "TRANSPARENT", "0", "#FFFFFF00"]):
            self.selected_hex = "TRANSPARENT"
        else:
            self.selected_hex = hex_code
        self.update_ui()

def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

class QualityDocApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.app_dir = get_app_dir()
        self.excel_path = os.path.join(self.app_dir, DEFAULT_EXCEL_NAME)
        self.wb = None

        self.title(f"AUTEL - Utilita pro Dokumentaci Kvality {APP_VERSION} (Autor: {APP_AUTHOR})")
        self.geometry("980x870")

        # Nastavení ikony okna
        icon_path = os.path.join(self.app_dir, "app_icon.ico")
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except Exception:
                pass

        # --- HLAVIČKA A LOGO ---
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(padx=20, pady=(12, 5), fill="x")

        # Načtení loga aplikace
        logo_path = os.path.join(self.app_dir, "app_logo.png")
        if os.path.exists(logo_path):
            try:
                pil_logo = Image.open(logo_path)
                ctk_logo = ctk.CTkImage(light_image=pil_logo, dark_image=pil_logo, size=(56, 56))
                lbl_logo = ctk.CTkLabel(header_frame, image=ctk_logo, text="")
                lbl_logo.pack(side="left", padx=(0, 15))
            except Exception:
                pass

        titles_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        titles_frame.pack(side="left", fill="both", expand=True)

        self.title_label = ctk.CTkLabel(
            titles_frame, 
            text=f"Utilita pro Tvorbu Dokumentace Kvality ({APP_VERSION})", 
            font=ctk.CTkFont(size=20, weight="bold"),
            anchor="w"
        )
        self.title_label.pack(anchor="w")

        self.sub_label = ctk.CTkLabel(
            titles_frame, 
            text=f"Autor: {APP_AUTHOR} (AUTEL, a.s.)  |  Automatické razítkování, konverze a kompletace PDF podle Excelu", 
            font=ctk.CTkFont(size=11, slant="italic"),
            text_color="#AAAAAA",
            anchor="w"
        )
        self.sub_label.pack(anchor="w", pady=(2, 0))

        # Tlačítko Návodu v hlavičce
        self.btn_manual = ctk.CTkButton(
            header_frame, 
            text="📖 Návod k použití", 
            command=self.open_manual_window,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#1F497D",
            hover_color="#143054",
            width=140,
            height=38
        )
        self.btn_manual.pack(side="right", padx=(10, 0))

        # --- TABVIEW PRO NASTAVENÍ ---
        self.tabview = ctk.CTkTabview(self, height=360)
        self.tabview.pack(padx=20, pady=5, fill="x")

        self.tabview.pack(padx=20, pady=5, fill="x")

        self.tab_dirs = self.tabview.add("Adresáře a Pravidla Zpracování")
        self.tab_stamp = self.tabview.add("Formát Razítka (Grafická paleta)")

        # === ZÁLOŽKA 1: Adresáře a Pravidla Zpracování ===
        f_paths = ctk.CTkFrame(self.tab_dirs, fg_color="transparent")
        f_paths.pack(fill="x", padx=5, pady=5)

        ctk.CTkLabel(f_paths, text="Zdrojový adresář:", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, padx=10, pady=4, sticky="w")
        self.entry_src = ctk.CTkEntry(f_paths, placeholder_text="Cesta ke zdrojovým souborům...")
        self.entry_src.grid(row=0, column=1, padx=10, pady=4, sticky="ew")
        ctk.CTkButton(f_paths, text="Procházet...", width=100, command=self.browse_source).grid(row=0, column=2, padx=10, pady=4)

        ctk.CTkLabel(f_paths, text="Cílový adresář:", font=ctk.CTkFont(size=12, weight="bold")).grid(row=1, column=0, padx=10, pady=4, sticky="w")
        self.entry_dst = ctk.CTkEntry(f_paths, placeholder_text="Cesta pro uložení zpracovaných PDF...")
        self.entry_dst.grid(row=1, column=1, padx=10, pady=4, sticky="ew")
        ctk.CTkButton(f_paths, text="Procházet...", width=100, command=self.browse_target).grid(row=1, column=2, padx=10, pady=4)

        ctk.CTkLabel(f_paths, text="Chybový adresář:", font=ctk.CTkFont(size=12)).grid(row=2, column=0, padx=10, pady=4, sticky="w")
        self.entry_err = ctk.CTkEntry(f_paths, placeholder_text="Cesta pro chybné soubory...")
        self.entry_err.grid(row=2, column=1, padx=10, pady=4, sticky="ew")
        ctk.CTkButton(f_paths, text="Procházet...", width=100, command=self.browse_error).grid(row=2, column=2, padx=10, pady=4)

        f_paths.grid_columnconfigure(1, weight=1)

        # Rámec pro pravidla zpracování přímo v 1. záložce
        f_rules = ctk.CTkFrame(self.tab_dirs, fg_color="transparent")
        f_rules.pack(fill="x", padx=5, pady=(0, 5))

        self.var_rotate = ctk.BooleanVar(value=True)
        self.chk_rotate = ctk.CTkCheckBox(f_rules, text="Otáčet stránky naležato (Landscape) na výšku (Portrait)", variable=self.var_rotate, command=self.on_setting_changed)
        self.chk_rotate.grid(row=0, column=0, padx=10, pady=4, sticky="w")

        self.var_keep_structure = ctk.BooleanVar(value=False)
        self.chk_keep_structure = ctk.CTkCheckBox(f_rules, text="Zachovat adresářovou strukturu v cílovém adresáři (kopírovat podadresáře)", variable=self.var_keep_structure, command=self.on_setting_changed)
        self.chk_keep_structure.grid(row=1, column=0, padx=10, pady=4, sticky="w")

        self.var_overwrite = ctk.BooleanVar(value=True)
        self.chk_overwrite = ctk.CTkCheckBox(f_rules, text="Přepsat existující soubory v cílovém adresáři", variable=self.var_overwrite, command=self.on_setting_changed)
        self.chk_overwrite.grid(row=2, column=0, padx=10, pady=4, sticky="w")

        self.var_merge = ctk.BooleanVar(value=False)
        self.chk_merge = ctk.CTkCheckBox(f_rules, text="Sloučit všechny dokumenty do jednoho výsledného PDF (Master PDF se záložkami)", variable=self.var_merge, command=self.on_setting_changed)
        self.chk_merge.grid(row=3, column=0, padx=10, pady=4, sticky="w")

        self.var_purge = ctk.BooleanVar(value=False)
        self.chk_purge = ctk.CTkCheckBox(f_rules, text="Smazat všechny soubory v cílovém adresáři před spuštěním", variable=self.var_purge, command=self.on_setting_changed)
        self.chk_purge.grid(row=4, column=0, padx=10, pady=4, sticky="w")

        # === ZÁLOŽKA 2: Formát Razítka (Grafická paleta) ===
        f_top = ctk.CTkFrame(self.tab_stamp, fg_color="transparent")
        f_top.pack(fill="x", padx=5, pady=5)

        ctk.CTkLabel(f_top, text="Velikost fontu (pt):", font=ctk.CTkFont(size=12)).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.entry_fontsize = ctk.CTkEntry(f_top, width=80)
        self.entry_fontsize.insert(0, "10")
        self.entry_fontsize.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(f_top, text="Průhlednost pozadí (%):", font=ctk.CTkFont(size=12)).grid(row=0, column=2, padx=(15, 5), pady=5, sticky="w")
        self.entry_opacity = ctk.CTkEntry(f_top, width=80)
        self.entry_opacity.insert(0, "100")
        self.entry_opacity.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        self.var_bold = ctk.BooleanVar(value=False)
        self.chk_bold = ctk.CTkCheckBox(f_top, text="Tučné (Bold)", variable=self.var_bold, command=self.on_setting_changed)
        self.chk_bold.grid(row=1, column=0, columnspan=2, padx=10, pady=8, sticky="w")

        self.var_italic = ctk.BooleanVar(value=False)
        self.chk_italic = ctk.CTkCheckBox(f_top, text="Kurzíva (Italic)", variable=self.var_italic, command=self.on_setting_changed)
        self.chk_italic.grid(row=1, column=2, padx=10, pady=8, sticky="w")

        self.var_underline = ctk.BooleanVar(value=False)
        self.chk_underline = ctk.CTkCheckBox(f_top, text="Podtržené", variable=self.var_underline, command=self.on_setting_changed)
        self.chk_underline.grid(row=1, column=3, padx=10, pady=8, sticky="w")

        f_palettes = ctk.CTkFrame(self.tab_stamp, fg_color="transparent")
        f_palettes.pack(fill="x", padx=5, pady=5)

        self.picker_text = GraphicalColorPicker(f_palettes, title="Barva textu razítka:", default_hex="#000000", on_change=self.on_setting_changed)
        self.picker_text.grid(row=0, column=0, padx=15, pady=5, sticky="nsew")

        self.picker_bg = GraphicalColorPicker(f_palettes, title="Barva pozadí razítka:", default_hex="#FFFFFF", is_bg=True, on_change=self.on_setting_changed)
        self.picker_bg.grid(row=0, column=1, padx=15, pady=5, sticky="nsew")

        f_palettes.grid_columnconfigure(0, weight=1)
        f_palettes.grid_columnconfigure(1, weight=1)

        # Navázání automatického ukládání při opuštění nebo zmáčknutí Enteru u textových polí
        for entry_widget in [self.entry_src, self.entry_dst, self.entry_err, self.entry_fontsize, self.entry_opacity]:
            entry_widget.bind("<FocusOut>", lambda e: self.on_setting_changed())
            entry_widget.bind("<Return>", lambda e: self.on_setting_changed())

        # --- TLAČÍTKA AKCÍ ---
        self.action_frame = ctk.CTkFrame(self)
        self.action_frame.pack(padx=20, pady=10, fill="x")

        self.btn_gen_excel = ctk.CTkButton(
            self.action_frame, 
            text="Vygenerovat / Obnovit Excel", 
            command=self.on_generate_excel,
            font=ctk.CTkFont(size=13, weight="bold"),
            height=42
        )
        self.btn_gen_excel.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        self.btn_open_excel = ctk.CTkButton(
            self.action_frame, 
            text="Otevřít Excel v MS Excelu", 
            command=self.on_open_excel,
            font=ctk.CTkFont(size=13, weight="bold"),
            height=42,
            fg_color="#1F497D",
            hover_color="#143054"
        )
        self.btn_open_excel.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        self.btn_run_process = ctk.CTkButton(
            self.action_frame, 
            text="Spustit razítkování podle Excelu", 
            command=self.on_run_process,
            font=ctk.CTkFont(size=13, weight="bold"),
            height=42,
            fg_color="#28a745",
            hover_color="#218838"
        )
        self.btn_run_process.grid(row=0, column=2, padx=10, pady=10, sticky="ew")

        self.action_frame.grid_columnconfigure(0, weight=1)
        self.action_frame.grid_columnconfigure(1, weight=1)
        self.action_frame.grid_columnconfigure(2, weight=1)

        # --- LOGY ---
        self.log_frame = ctk.CTkFrame(self)
        self.log_frame.pack(padx=20, pady=(5, 15), fill="both", expand=True)

        self.lbl_status = ctk.CTkLabel(
            self.log_frame, 
            text=f"Stav Excelu: Soubor '{DEFAULT_EXCEL_NAME}' nepřipojen.",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.lbl_status.pack(padx=10, pady=(10, 5), anchor="w")

        self.log_text = ctk.CTkTextbox(self.log_frame, font=ctk.CTkFont(family="Consolas", size=11))
        self.log_text.pack(padx=10, pady=(0, 10), fill="both", expand=True)

        self.check_initial_excel()

    def log(self, message):
        self.log_text.insert("end", f"{message}\n")
        self.log_text.see("end")

    def get_initial_dir(self, entry_widget):
        path = entry_widget.get().strip()
        if path and os.path.exists(path):
            return path
        elif path:
            parent = os.path.dirname(path)
            if os.path.exists(parent):
                return parent
        return self.app_dir

    def on_setting_changed(self, *args):
        """Automaticky a okamžitě uloží veškerá nastavení z GUI do záložky AUTEL v Excelu při jakékoliv změně."""
        self.update_excel_config_from_gui()

    def update_excel_config_from_gui(self):
        """Uloží aktuální cesty a nastavení z GUI do záložky AUTEL v Excelu (přímo nebo přes COM)."""
        if not os.path.exists(self.excel_path):
            return
        try:
            cfg_dict = self.get_gui_config_dict()
            if self.wb is None:
                self.wb = safe_load_workbook(self.excel_path)
            update_autel_config(self.wb, cfg_dict)
            saved, save_msg = safe_save_workbook(self.wb, self.excel_path, config_dict=cfg_dict)
            if saved:
                self.lbl_status.configure(text=f"Stav Excelu: Nastavení uloženo do záložky AUTEL ({DEFAULT_EXCEL_NAME}).")
                self.log(f"Změna nastavení automaticky uložena do záložky AUTEL.")
        except Exception as e:
            self.log(f"Varování při zápisu nastavení do Excelu: {e}")

    def browse_source(self):
        init_dir = self.get_initial_dir(self.entry_src)
        d = filedialog.askdirectory(title="Vyberte zdrojovou složku", initialdir=init_dir)
        if d:
            self.entry_src.delete(0, "end")
            self.entry_src.insert(0, d)

            # Zachovat stávající Cíl a Chyby, pokud jsou v GUI již nastaveny!
            if not self.entry_dst.get().strip():
                self.entry_dst.insert(0, os.path.join(d, "Cíl"))
            if not self.entry_err.get().strip():
                self.entry_err.insert(0, os.path.join(d, "Chyby"))
            
            self.update_excel_config_from_gui()
            self.log(f"Změna zdrojového adresáře: {d}")

    def browse_target(self):
        init_dir = self.get_initial_dir(self.entry_dst)
        d = filedialog.askdirectory(title="Vyberte cílovou složku", initialdir=init_dir)
        if d:
            self.entry_dst.delete(0, "end")
            self.entry_dst.insert(0, d)
            self.update_excel_config_from_gui()

    def browse_error(self):
        init_dir = self.get_initial_dir(self.entry_err)
        d = filedialog.askdirectory(title="Vyberte složku pro chyby", initialdir=init_dir)
        if d:
            self.entry_err.delete(0, "end")
            self.entry_err.insert(0, d)
            self.update_excel_config_from_gui()

    def get_gui_config_dict(self):
        tc = self.picker_text.get_hex()
        bg_hex = self.picker_bg.get_hex()
        bg = "#FFFFFF" if bg_hex == "TRANSPARENT" else bg_hex
        op = "0" if bg_hex == "TRANSPARENT" else self.entry_opacity.get().strip()

        return {
            "Zdroj": self.entry_src.get().strip(),
            "Cíl": self.entry_dst.get().strip(),
            "Chybné soubory": self.entry_err.get().strip(),
            "Přepsat existující soubory": "ANO" if self.var_overwrite.get() else "NE",
            "Smazat cílový adresář": "ANO" if self.var_purge.get() else "NE",
            "Otáčet stránky na výšku": "ANO" if self.var_rotate.get() else "NE",
            "Zachovat adresářovou strukturu": "ANO" if self.var_keep_structure.get() else "NE",
            "Kompletovat do jednoho PDF": "ANO" if self.var_merge.get() else "NE",
            "Velikost fontu razítka": self.entry_fontsize.get().strip(),
            "Tučné razítko": "ANO" if self.var_bold.get() else "NE",
            "Kurzíva razítka": "ANO" if self.var_italic.get() else "NE",
            "Podtržené razítko": "ANO" if self.var_underline.get() else "NE",
            "Barva fontu razítka": tc,
            "Barva pozadí razítka": bg,
            "Průhlednost pozadí (%)": op,
            "Detekovat prázdné stránky": "ANO"
        }

    def check_initial_excel(self):
        if os.path.exists(self.excel_path):
            self.log(f"Nalezen stávající soubor Excel vedle utility: {self.excel_path}")
            try:
                wb = safe_load_workbook(self.excel_path)
                config = read_config_from_excel(wb)
                
                if config.get("Zdroj"):
                    self.entry_src.delete(0, "end")
                    self.entry_src.insert(0, config.get("Zdroj"))
                if config.get("Cíl"):
                    self.entry_dst.delete(0, "end")
                    self.entry_dst.insert(0, config.get("Cíl"))
                if config.get("Chybné soubory"):
                    self.entry_err.delete(0, "end")
                    self.entry_err.insert(0, config.get("Chybné soubory"))
                if config.get("Velikost fontu razítka"):
                    self.entry_fontsize.delete(0, "end")
                    self.entry_fontsize.insert(0, config.get("Velikost fontu razítka"))
                if config.get("Průhlednost pozadí (%)"):
                    self.entry_opacity.delete(0, "end")
                    self.entry_opacity.insert(0, config.get("Průhlednost pozadí (%)"))

                if config.get("Barva fontu razítka"):
                    self.picker_text.set_hex(config.get("Barva fontu razítka"))
                if config.get("Barva pozadí razítka"):
                    bg_val = config.get("Barva pozadí razítka")
                    if config.get("Průhlednost pozadí (%)") == "0":
                        bg_val = "TRANSPARENT"
                    self.picker_bg.set_hex(bg_val)

                self.var_bold.set(config.get("Tučné razítko", "NE").upper() == "ANO")
                self.var_italic.set(config.get("Kurzíva razítka", "NE").upper() == "ANO")
                self.var_underline.set(config.get("Podtržené razítko", "NE").upper() == "ANO")

                self.var_rotate.set(config.get("Otáčet stránky na výšku", "ANO").upper() == "ANO")
                self.var_keep_structure.set(config.get("Zachovat adresářovou strukturu", "NE").upper() == "ANO")
                self.var_overwrite.set(config.get("Přepsat existující soubory", "ANO").upper() == "ANO")
                self.var_merge.set(config.get("Kompletovat do jednoho PDF", "NE").upper() == "ANO")
                self.var_purge.set(config.get("Smazat cílový adresář", "NE").upper() == "ANO")
                    
                self.lbl_status.configure(text=f"Stav Excelu: Nalezen a načten '{DEFAULT_EXCEL_NAME}' vedle utility.")
                self.log("Automaticky otevírám soubor Excel v aplikaci...")
                self.on_open_excel()
            except Exception as e:
                self.log(f"Varování při načítání stávajícího Excelu: {e}")
        else:
            self.log(f"Soubor '{DEFAULT_EXCEL_NAME}' vedle utility neexistuje.")
            self.log("Vyberte prosím Zdrojový a Cílový adresář a klikněte na 'Vygenerovat / Obnovit Excel'.")

    def on_generate_excel(self):
        source_dir = self.entry_src.get().strip()
        target_dir = self.entry_dst.get().strip()
        error_dir = self.entry_err.get().strip()

        if not source_dir or not os.path.exists(source_dir):
            messagebox.showerror("Chyba", "Prosím vyberte platný Zdrojový adresář.")
            return

        if not target_dir:
            target_dir = os.path.join(source_dir, "Cíl")
            self.entry_dst.delete(0, "end")
            self.entry_dst.insert(0, target_dir)

        if not error_dir:
            error_dir = os.path.join(source_dir, "Chyby")
            self.entry_err.delete(0, "end")
            self.entry_err.insert(0, error_dir)

        self.log(f"Generuji Excel vedle utility pro složku: {source_dir}")
        self.log("Skenuji podadresáře a vyčítám strany (vylučuji Cíl a Chyby)...")

        try:
            wb, path, created = get_or_create_excel(self.excel_path, source_dir)
            self.wb = wb
            
            cfg_dict = self.get_gui_config_dict()
            update_autel_config(self.wb, cfg_dict)

            # Načíst stávající Označení z Excelu, aby se při obnovení neztratily ručně zadané hodnoty!
            existing_map = {}
            if os.path.exists(self.excel_path):
                try:
                    com_cfg, com_entries = read_open_excel_via_com(self.excel_path)
                    if com_entries:
                        for e in com_entries:
                            p = e.get('cesta', '').strip().lower()
                            ozn = e.get('oznaceni', '').strip()
                            if p and ozn:
                                existing_map[p] = ozn
                                existing_map[os.path.abspath(p).lower()] = ozn
                    else:
                        ws_seznam = self.wb["Seznam"]
                        for r in range(2, ws_seznam.max_row + 1):
                            ozn = ws_seznam.cell(r, 1).value
                            p = ws_seznam.cell(r, 2).value
                            if p and ozn:
                                existing_map[str(p).strip().lower()] = str(ozn).strip()
                                existing_map[os.path.abspath(str(p)).lower()] = str(ozn).strip()
                except Exception as e_map:
                    self.log(f"Varování při načítání stávajících označení: {e_map}")

            file_entries = scan_directory(
                source_dir, 
                check_blank_pages=True, 
                target_dir=target_dir,
                error_dir=error_dir,
                existing_map=existing_map
            )
            self.log(f"Naskenováno {len(file_entries)} souborů podkladů.")
            
            write_files_to_seznam(self.wb, file_entries)
            
            saved, save_msg = safe_save_workbook(self.wb, self.excel_path, file_entries=file_entries, config_dict=cfg_dict)
            if saved:
                if "COM" in save_msg or "otvír" in save_msg or "otevřen" in save_msg:
                    self.log("Nová data byla zapsána přímo do vašeho otevřeného okna MS Excel!")
                else:
                    self.log(f"Excel byl vytvořen/obnoven: {self.excel_path}")
            else:
                self.log(f"UPOZORNĚNÍ: Soubor '{DEFAULT_EXCEL_NAME}' je otevřen v MS Excelu. Všechna data byla naskenována. Pro zobrazení nově naskenovaných dat v Excelu ukončete úpravu buňky v Excelu a klikněte na 'Obnovit Excel'.")

            self.lbl_status.configure(text=f"Stav Excelu: Vygenerován '{DEFAULT_EXCEL_NAME}' ({len(file_entries)} souborů).")
            self.on_open_excel()

        except Exception as e:
            self.log(f"CHYBA při generování Excelu: {e}")
            messagebox.showerror("Chyba", f"Nepodařilo se vygenerovat Excel: {e}")

    def on_open_excel(self):
        if not os.path.exists(self.excel_path):
            messagebox.showwarning("Varování", f"Soubor Excel '{DEFAULT_EXCEL_NAME}' ještě neexistuje. Nejprve jej vygenerujte.")
            return

        try:
            if sys.platform == "win32":
                os.startfile(self.excel_path)
            else:
                subprocess.call(["open", self.excel_path])
        except Exception as e:
            self.log(f"CHYBA při otevírání Excelu: {e}")

    def on_run_process(self):
        if not os.path.exists(self.excel_path):
            messagebox.showwarning("Varování", f"Soubor Excel '{DEFAULT_EXCEL_NAME}' neexistuje.")
            return

        # 1. Uložit aktuální cestu Cíle a Chyb z GUI do Excelu
        self.update_excel_config_from_gui()

        self.log("Zahajuji razítkování a kompletaci PDF podle Excelu...")
        try:
            wb = safe_load_workbook(self.excel_path)

            processed, errors = process_pdfs_from_excel(wb, excel_path=self.excel_path, log_callback=self.log)
            
    def open_manual_window(self):
        """Otevře samostatné okno s interaktivním návodem přímo v aplikaci."""
        ManualWindow(self)

    def open_pdf_manual(self):
        """Otevře oficiální PDF příručku 'Uzivatelska_Prirucka_Utilita_Dokumentace_Kvality.pdf'."""
        pdf_path = os.path.join(self.app_dir, MANUAL_PDF_NAME)
        if not os.path.exists(pdf_path):
            messagebox.showwarning("Varování", f"Soubor příručky '{MANUAL_PDF_NAME}' nebyl nalezen v adresáři aplikace.")
            return
        try:
            if sys.platform == "win32":
                os.startfile(pdf_path)
            else:
                subprocess.call(["open", pdf_path])
        except Exception as e:
            self.log(f"CHYBA při otevírání PDF příručky: {e}")

def run_app():
    app = QualityDocApp()
    app.mainloop()

if __name__ == "__main__":
    run_app()
