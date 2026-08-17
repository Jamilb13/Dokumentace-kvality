import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEFAULT_EXCEL_NAME = "Dokumentace_Kvality.xlsx"

BLUE_TAB_COLOR = "1F497D"
HEADER_FILL_COLOR = "1F497D"
HEADER_FONT_COLOR = "FFFFFF"

def copy_file_windows_native(src, dst):
    """
    Zkopíruje soubor pomocí nativního Windows Kernel32 API.
    Spolehlivě obchází standardní Python C-runtime fopen zámky u MS Excelu a OneDrive.
    """
    import ctypes
    res = ctypes.windll.kernel32.CopyFileW(str(src), str(dst), False)
    if not res:
        err = ctypes.GetLastError()
        raise PermissionError(f"Windows API CopyFileW selhalo s kódem {err}")
    return True

def safe_load_workbook(excel_path, data_only=False):
    """
    Bezpečně načte sešit openpyxl z excel_path.
    Pokud je soubor otevřen v MS Excelu / OneDrive (PermissionError),
    vytvoří v %TEMP% dočasnou stínovou kopii pomocí nativního Windows Kernel32 API.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Soubor neexistuje: {excel_path}")
        
    try:
        return openpyxl.load_workbook(excel_path, data_only=data_only)
    except PermissionError:
        print(f"Soubor {excel_path} je uzamčen v MS Excelu/OneDrive. Načítám přes nativní Windows Kernel32 API...")
        import tempfile
        
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_read_{os.path.basename(excel_path)}")
        try:
            copy_file_windows_native(excel_path, temp_file)
            wb = openpyxl.load_workbook(temp_file, data_only=data_only)
            return wb
        except Exception as e:
            raise PermissionError(f"Soubor '{os.path.basename(excel_path)}' je zablokován v MS Excelu a nelze jej přečíst: {e}")

def read_open_excel_via_com(excel_path):
    """
    Vyčte konfiguraci AUTEL i Seznam souborů ŽIVĚ z otevřeného okna MS Excel přes COM API.
    Vrací (config_dict, file_entries) nebo (None, None) pokud MS Excel není spuštěn s daným souborem.
    """
    try:
        import win32com.client
        try:
            excel_app = win32com.client.GetObject(Class="Excel.Application")
        except Exception:
            return None, None
            
        abs_path = os.path.abspath(excel_path).lower()
        target_wb = None
        for wb in excel_app.Workbooks:
            if wb.FullName.lower() == abs_path or wb.Name.lower() == os.path.basename(abs_path).lower():
                target_wb = wb
                break
                
        if not target_wb:
            return None, None
            
        print(f"Čtu konfiguraci i seznam ŽIVĚ z MS Excel přes COM API: {target_wb.Name}")
        
        # 1. Čtení AUTEL
        config = {}
        try:
            ws_autel = target_wb.Worksheets("AUTEL")
            used_rows = ws_autel.UsedRange.Rows.Count
            for r in range(2, used_rows + 1):
                k = ws_autel.Cells(r, 1).Value
                v = ws_autel.Cells(r, 2).Value
                if k:
                    config[str(k).strip()] = str(v).strip() if v is not None else ""
        except Exception as e:
            print(f"Varování při COM čtení AUTEL: {e}")

        # 2. Čtení Seznam
        file_entries = []
        try:
            ws_seznam = target_wb.Worksheets("Seznam")
            used_rows = ws_seznam.UsedRange.Rows.Count
            for r in range(2, used_rows + 1):
                oznaceni = ws_seznam.Cells(r, 1).Value
                cesta = ws_seznam.Cells(r, 2).Value
                nazev = ws_seznam.Cells(r, 3).Value
                strany = ws_seznam.Cells(r, 4).Value
                zapis = ws_seznam.Cells(r, 5).Value
                novy_nazev = ws_seznam.Cells(r, 6).Value
                
                if cesta or nazev or oznaceni:
                    file_entries.append({
                        'row_idx': r,
                        'oznaceni': str(oznaceni).strip() if oznaceni is not None else "",
                        'cesta': str(cesta).strip() if cesta is not None else "",
                        'nazev': str(nazev).strip() if nazev is not None else "",
                        'pocet_stran': int(strany) if strany is not None and str(strany).isdigit() else 0,
                        'zapis': str(zapis).strip() if zapis is not None else "",
                        'novy_nazev': str(novy_nazev).strip() if novy_nazev is not None else ""
                    })
        except Exception as e:
            print(f"Varování při COM čtení Seznam: {e}")

        return config, file_entries
    except Exception as e:
        print(f"COM čtení selhalo: {e}")
        return None, None

def write_to_open_excel_via_com(excel_path, file_entries=None, config_dict=None, only_update_status=False):
    """
    Pokud je sešit otevřen v běžící aplikaci MS Excel (Windows COM API),
    zapíše data přímo do živé instance MS Excelu a uloží sešit.
    Obsahuje vyhledání v ROT a retries pro případ dočasné zaneprázdněnosti MS Excelu.
    """
    import time
    import win32com.client

    abs_path = os.path.abspath(excel_path)
    abs_path_lower = abs_path.lower()
    base_name_lower = os.path.basename(abs_path).lower()

    max_attempts = 4
    for attempt in range(max_attempts):
        target_wb = None
        
        # 1. Zkusit přímo GetObject s celou cestou (získání otevřeného sešitu přímo z ROT)
        try:
            wb_obj = win32com.client.GetObject(Pathname=abs_path)
            if wb_obj:
                target_wb = wb_obj
        except Exception:
            pass

        # 2. Pokud selhalo, zkusit projít běžící aplikaci Excel
        if not target_wb:
            try:
                excel_app = win32com.client.GetObject(Class="Excel.Application")
                if excel_app:
                    for wb in excel_app.Workbooks:
                        if (wb.FullName and wb.FullName.lower() == abs_path_lower) or \
                           (wb.Name and wb.Name.lower() == base_name_lower):
                            target_wb = wb
                            break
            except Exception:
                pass

        if not target_wb:
            if attempt < max_attempts - 1:
                time.sleep(0.3)
                continue
            return False

        try:
            print(f"Připojuji se přímo k otevřenému sešitu v MS Excel: {target_wb.Name}")
            
            # 1. Aktualizovat konfiguraci na záložce AUTEL
            if config_dict:
                try:
                    ws_autel = target_wb.Worksheets("AUTEL")
                    used_rows = ws_autel.UsedRange.Rows.Count
                    key_to_row = {}
                    for r in range(2, used_rows + 1):
                        k = ws_autel.Cells(r, 1).Value
                        if k:
                            key_to_row[str(k).strip()] = r
                            
                    for k, v in config_dict.items():
                        if k in key_to_row:
                            ws_autel.Cells(key_to_row[k], 2).Value = str(v)
                        else:
                            new_r = ws_autel.UsedRange.Rows.Count + 1
                            ws_autel.Cells(new_r, 1).Value = str(k)
                            ws_autel.Cells(new_r, 2).Value = str(v)
                except Exception as e:
                    print(f"Varování při zápisu konfigurace přes COM: {e}")

            # 2. Zapsat soubory na záložku Seznam
            if file_entries is not None:
                try:
                    ws_seznam = target_wb.Worksheets("Seznam")
                except Exception:
                    ws_seznam = target_wb.Worksheets.Add()
                    ws_seznam.Name = "Seznam"
                    
                if only_update_status:
                    # Pouze aktualizace stavu a nového názvu u zpracovávaných řádků – BEZ mazání buněk!
                    for entry in file_entries:
                        r = entry.get('row_idx')
                        if r and r >= 2:
                            status = str(entry.get('zapis', ''))
                            novy_n = str(entry.get('novy_nazev', ''))
                            ws_seznam.Cells(r, 5).Value = status
                            ws_seznam.Cells(r, 6).Value = novy_n

                            st_upper = status.upper()
                            if "CHYBA" in st_upper or entry.get('pocet_stran', 0) == 0:
                                ws_seznam.Range(ws_seznam.Cells(r, 1), ws_seznam.Cells(r, 6)).Interior.Color = 13553407 # Světle červená
                            elif "VAROVÁNÍ" in st_upper or "VAROVANI" in st_upper:
                                ws_seznam.Range(ws_seznam.Cells(r, 1), ws_seznam.Cells(r, 6)).Interior.Color = 10283519 # Světle žlutá
                            elif "OK" in st_upper:
                                ws_seznam.Range(ws_seznam.Cells(r, 1), ws_seznam.Cells(r, 6)).Interior.Color = 13561542 # Světle zelená
                else:
                    # Plné vygenerování/obnovení seznamu souborů
                    for idx, entry in enumerate(file_entries, start=2):
                        ws_seznam.Cells(idx, 1).Value = str(entry.get('oznaceni', ''))
                        ws_seznam.Cells(idx, 2).Value = str(entry.get('cesta', ''))
                        ws_seznam.Cells(idx, 3).Value = str(entry.get('nazev', ''))
                        ws_seznam.Cells(idx, 4).Value = int(entry.get('pocet_stran', 0))
                        ws_seznam.Cells(idx, 5).Value = str(entry.get('zapis', ''))
                        ws_seznam.Cells(idx, 6).Value = str(entry.get('novy_nazev', ''))

                        status = str(entry.get('zapis', '')).upper()
                        if "CHYBA" in status or entry.get('pocet_stran', 0) == 0:
                            ws_seznam.Range(ws_seznam.Cells(idx, 1), ws_seznam.Cells(idx, 6)).Interior.Color = 13553407 # Světle červená
                        elif "VAROVÁNÍ" in status or "VAROVANI" in status:
                            ws_seznam.Range(ws_seznam.Cells(idx, 1), ws_seznam.Cells(idx, 6)).Interior.Color = 10283519 # Světle žlutá
                        elif "OK" in status:
                            ws_seznam.Range(ws_seznam.Cells(idx, 1), ws_seznam.Cells(idx, 6)).Interior.Color = 13561542 # Světle zelená

                    used_rows = ws_seznam.UsedRange.Rows.Count
                    new_count = len(file_entries) + 1
                    if used_rows > new_count:
                        # Vyčistit pouze nadbytečné řádky pod nově naskenovaným seznamem
                        ws_seznam.Range(ws_seznam.Cells(new_count + 1, 1), ws_seznam.Cells(used_rows + 5, 6)).ClearContents()
                        ws_seznam.Range(ws_seznam.Cells(new_count + 1, 1), ws_seznam.Cells(used_rows + 5, 6)).Interior.ColorIndex = 0

            target_wb.Save()
            return True

        except Exception as e:
            print(f"COM pokus {attempt+1}/{max_attempts} selhal: {e}")
            if attempt < max_attempts - 1:
                time.sleep(0.4)
            else:
                return False

    return False

def safe_save_workbook(wb, excel_path, file_entries=None, config_dict=None, only_update_status=False):
    """
    Uloží sešit. Pokud je otevřen v MS Excelu (PermissionError),
    využije COM rozhraní k přímému zápisu do otevřeného okna MS Excelu.
    """
    try:
        wb.save(excel_path)
        return True, ""
    except PermissionError:
        print("Excel soubor je otevřen. Zkouším zápis přes COM API rozhraní MS Excel...")
        success = write_to_open_excel_via_com(excel_path, file_entries=file_entries, config_dict=config_dict, only_update_status=only_update_status)
        if success:
            return True, "Zapsáno přímo do otevřeného okna MS Excel."
        else:
            err_msg = (
                f"Soubor '{os.path.basename(excel_path)}' je otevřen v programu MS Excel!\n\n"
                "Uložení stavů do Excelu selhalo. PDF soubory v cílové složce však byly vygenerovány."
            )
            return False, err_msg
    except Exception as e:
        return False, f"Nepodařilo se uložit Excel: {e}"

def get_or_create_excel(excel_path=None, source_dir=None):
    if excel_path is None:
        excel_path = DEFAULT_EXCEL_NAME
        
    if os.path.exists(excel_path):
        try:
            wb = safe_load_workbook(excel_path)
            return wb, excel_path, False
        except Exception:
            pass
    
    wb = openpyxl.Workbook()
    
    ws_autel = wb.active
    ws_autel.title = "AUTEL"
    ws_autel.sheet_properties.tabColor = BLUE_TAB_COLOR
    
    setup_autel_sheet(ws_autel, source_dir)
    
    ws_seznam = wb.create_sheet(title="Seznam")
    setup_seznam_sheet(ws_seznam)
    
    safe_save_workbook(wb, excel_path)
    return wb, excel_path, True

def setup_autel_sheet(ws, source_dir=None, config_dict=None):
    cfg = config_dict or {}
    
    headers = [
        ("Parametr", "Hodnota", "Popis"),
        ("Zdroj", cfg.get("Zdroj", source_dir if source_dir else ""), "Zdrojová složka s PDF a podklady"),
        ("Cíl", cfg.get("Cíl", os.path.join(source_dir, "Cíl") if source_dir else ""), "Cílová složka pro zpracovaná PDF"),
        ("Chybné soubory", cfg.get("Chybné soubory", os.path.join(source_dir, "Chyby") if source_dir else ""), "Složka pro nečitelná / chybná PDF"),
        ("Přepsat existující soubory", cfg.get("Přepsat existující soubory", "ANO"), "Přepsat soubory v cílové složce (ANO/NE)"),
        ("Smazat cílový adresář", cfg.get("Smazat cílový adresář", "NE"), "Vyčistit cílovou složku před spuštěním (ANO/NE)"),
        ("Otáčet stránky na výšku", cfg.get("Otáčet stránky na výšku", "ANO"), "Otočit landscape stránky na portrait před razítkováním (ANO/NE)"),
        ("Zachovat adresářovou strukturu", cfg.get("Zachovat adresářovou strukturu", "NE"), "Vytvářet podadresáře v cíli podle struktury zdrojové složky (ANO/NE)"),
        ("Kompletovat do jednoho PDF", cfg.get("Kompletovat do jednoho PDF", "NE"), "Sloučit všechna zpracovaná PDF do jednoho Master PDF (ANO/NE)"),
        ("Velikost fontu razítka", cfg.get("Velikost fontu razítka", "10"), "Velikost písma v pt (např. 10)"),
        ("Tučné razítko", cfg.get("Tučné razítko", "NE"), "Použít tučný font (ANO/NE)"),
        ("Kurzíva razítka", cfg.get("Kurzíva razítka", "NE"), "Použít kurzívu (ANO/NE)"),
        ("Podtržené razítko", cfg.get("Podtržené razítko", "NE"), "Podtrhnout text razítka (ANO/NE)"),
        ("Barva fontu razítka", cfg.get("Barva fontu razítka", "#000000"), "HEX kód barvy textu (např. #000000)"),
        ("Barva pozadí razítka", cfg.get("Barva pozadí razítka", "#FFFFFF"), "HEX kód pozadí razítka (např. #FFFFFF)"),
        ("Průhlednost pozadí (%)", cfg.get("Průhlednost pozadí (%)", "100"), "Průhlednost pozadí razítka v % (100 = neprůhledné, 0 = průhledné)"),
        ("Detekovat prázdné stránky", cfg.get("Detekovat prázdné stránky", "ANO"), "Upozornit na prázdné stránky ve zdrojových PDF (ANO/NE)"),
    ]
    
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for row_idx, row_data in enumerate(headers, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if col_idx == 1:
                    cell.font = Font(bold=True)
                elif col_idx == 2:
                    cell.alignment = Alignment(horizontal="left" if isinstance(val, str) and (":\\" in val or "/" in val) else "center")

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50

def update_autel_config(wb, config_dict):
    ws = wb["AUTEL"]
    key_to_row = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k:
            key_to_row[str(k).strip()] = r
            
    for k, v in config_dict.items():
        if k in key_to_row:
            ws.cell(row=key_to_row[k], column=2, value=str(v))
        else:
            new_r = ws.max_row + 1
            ws.cell(row=new_r, column=1, value=k)
            ws.cell(row=new_r, column=2, value=str(v))

def setup_seznam_sheet(ws):
    headers = [
        "Označení", "Úplná cesta", "Název souboru", "Počet stran", "Zápis", "Nový název souboru"
    ]
    
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    widths = [15, 60, 35, 12, 15, 45]
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    for c, w in zip(cols, widths):
        ws.column_dimensions[c].width = w

def read_config_from_excel(wb):
    ws = wb["AUTEL"]
    config = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, column=1).value
        val = ws.cell(row, column=2).value
        if key:
            config[str(key).strip()] = str(val).strip() if val is not None else ""
    return config

def write_files_to_seznam(wb, file_entries, only_update_status=False):
    ws = wb["Seznam"]
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    if only_update_status:
        for entry in file_entries:
            r = entry.get('row_idx')
            if r and r >= 2 and r <= ws.max_row:
                ws.cell(r, 5, value=entry.get('zapis', ''))
                ws.cell(r, 6, value=entry.get('novy_nazev', ''))
                status = str(entry.get('zapis', '')).upper()
                fill_to_apply = None
                if "CHYBA" in status or entry.get('pocet_stran', 0) == 0:
                    fill_to_apply = red_fill
                elif "VAROVÁNÍ" in status or "VAROVANI" in status:
                    fill_to_apply = yellow_fill
                elif "OK" in status:
                    fill_to_apply = green_fill
                if fill_to_apply:
                    for col_idx in range(1, 7):
                        ws.cell(r, col_idx).fill = fill_to_apply
        return

    for idx, entry in enumerate(file_entries, start=2):
        row_vals = [
            entry.get('oznaceni', ''),
            entry.get('cesta', ''),
            entry.get('nazev', ''),
            entry.get('pocet_stran', 0),
            entry.get('zapis', ''),
            entry.get('novy_nazev', '')
        ]
        
        status = str(entry.get('zapis', '')).upper()
        fill_to_apply = None
        if "CHYBA" in status or entry.get('pocet_stran', 0) == 0:
            fill_to_apply = red_fill
        elif "VAROVÁNÍ" in status or "VAROVANI" in status:
            fill_to_apply = yellow_fill
        elif "OK" in status:
            fill_to_apply = green_fill
            
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = thin_border
            if fill_to_apply:
                cell.fill = fill_to_apply
            if col_idx in [1, 4, 5]:
                cell.alignment = Alignment(horizontal="center")
